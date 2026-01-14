import io
import pandas as pd
from openpyxl.utils import get_column_letter

STATUS_KO={"success":"성공","failed":"실패","skipped":"제외"}
PARSEMODE_KO={"requests":"자동","playwright":"브라우저","skipped":"제외","unknown":"알수없음","exception":"오류"}
SITE_KO={"KYobo":"교보문고","YES24":"YES24","ALADIN":"알라딘","YPBOOKS":"영풍문고"}
COLUMN_KO={"site":"서점","url":"상품 URL","status":"처리상태","isbn":"ISBN","title":"도서명","author":"저자",
           "publisher":"출판사","list_price":"정가","sale_price":"판매가","product_id":"상품ID",
           "parse_mode":"처리방식","error":"오류","note":"비고"}

def to_xlsx_bytes(df_raw: pd.DataFrame) -> bytes:
    df=df_raw.copy()
    if "note" not in df.columns:
        df["note"]=""

    if "site" in df.columns:
        df["site"]=df["site"].map(SITE_KO).fillna(df["site"])
    if "status" in df.columns:
        df["status"]=df["status"].map(STATUS_KO).fillna(df["status"])
    if "parse_mode" in df.columns:
        df["parse_mode"]=df["parse_mode"].map(PARSEMODE_KO).fillna(df["parse_mode"])

    # --- Excel에서는 '처리상태' 컬럼 제거(요청 반영) ---
    if "status" in df.columns:
        df = df.drop(columns=["status"])

    df=df.rename(columns=COLUMN_KO)

    # --- URL은 맨 오른쪽으로 ---
    # 기본 컬럼 순서(요청: 처리상태 제외, URL 맨 오른쪽)
    preferred=["서점","ISBN","도서명","저자","출판사","정가","판매가","비고","상품ID","처리방식","오류"]
    cols=[c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred and c != "상품 URL"]
    if "상품 URL" in df.columns:
        cols = cols + ["상품 URL"]
    df=df[cols]

    out=io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="결과")
        ws=writer.sheets["결과"]
        for i,col in enumerate(df.columns, start=1):
            max_len=max([len(str(col))]+[len(str(x)) for x in df[col].astype(str).head(200)])
            ws.column_dimensions[get_column_letter(i)].width=min(max(10, max_len+2), 60)

        won_format='#,##0"원"'
        for money_col in ["정가","판매가"]:
            if money_col in df.columns:
                idx=df.columns.get_loc(money_col)+1
                for r in range(2, ws.max_row+1):
                    cell=ws.cell(row=r, column=idx)
                    if isinstance(cell.value,(int,float)) and cell.value is not None:
                        cell.number_format=won_format
    return out.getvalue()
